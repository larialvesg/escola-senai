import openpyxl
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, ListAPIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.shortcuts import render
from .models import Ambiente
from app_empresa.api.serializers import AmbienteSerializer
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, ListAPIView
from django.http import HttpResponse
from .forms import (
    ExcelUploadFormArea, ExcelUploadFormAmbiente, ExcelUploadFormPatrimonio, ExcelUploadFormFuncionario
)
from .models import Area, Ambiente, Funcionario, Patrimonio

def upload_area(request):
    if request.method == 'POST':
        form = ExcelUploadFormArea(request.POST, request.FILES)
        if form.is_valid():
            wb = openpyxl.load_workbook(request.FILES['file'])
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                Area.objects.create(area=row[0])
    else:
        form = ExcelUploadFormArea()
    return render(request, 'upload_excel.html', {'form': form, 'titulo': 'Upload de Área'})

def upload_ambiente(request):
    if request.method == 'POST':
        form = ExcelUploadFormAmbiente(request.POST, request.FILES)
        if form.is_valid():
            wb = openpyxl.load_workbook(request.FILES['file'])
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                Ambiente.objects.create(sig=row[0], descricao=row[1], sn=row[2], responsavel=row[3])
    else:
        form = ExcelUploadFormAmbiente()
    return render(request, 'upload_excel.html', {'form': form, 'titulo': 'Upload de Ambiente'})


def upload_funcionario(request):
    if request.method == 'POST':
        form = ExcelUploadFormFuncionario(request.POST, request.FILES)
        if form.is_valid():
            wb = openpyxl.load_workbook(request.FILES['file'])
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                Funcionario.objects.create(sn=row[0], nome=row[1], email=row[2], cargo=row[3], area=row[4])
    else:
        form = ExcelUploadFormFuncionario()
    return render(request, 'upload_excel.html', {'form': form, 'titulo': 'Upload de Funcionario'})



def upload_patrimonio(request):
    if request.method == 'POST':
        form = ExcelUploadFormPatrimonio(request.POST, request.FILES)
        if form.is_valid():
            wb = openpyxl.load_workbook(request.FILES['file'])
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                Patrimonio.objects.create(ni=row[1], descricao=row[2], localizacao=row[0])
    else:
        form = ExcelUploadFormPatrimonio()
    return render(request, 'upload_excel.html', {'form': form, 'titulo': 'Upload de Patrimônio'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def buscar_ambiente_por_id(request, idSearch):
    try:
        ambiente = Ambiente.objects.get(id=idSearch)
        serializer = AmbienteSerializer(ambiente)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Ambiente.DoesNotExist:
        return Response({'error': 'Ambiente não encontrado'}, status=status.HTTP_404_NOT_FOUND)
    